from openpyxl import load_workbook

def write_overall_report(
        result,
        target_month,
        template_path,
        output_path,
):
    workbook = load_workbook(template_path)

    sheet = workbook.active

    sheet["A1"] = f"【{target_month}】全体月次報告書"
    sheet["B3"] = target_month
    sheet["B4"] = "ALL"
    sheet["D4"] = "全体"

    sheet["B7"] = result["sales_amount_tax_ex"]
    sheet["D7"] = result["tax_amount"]
    sheet["F7"] = result["sales_amount_tax_in"]

    sheet["B10"] = result["customer_count"]
    sheet["D10"] = result["average_spend"]
    sheet["F10"] = result["tax_rate"]

    sheet["B13"] = result["cost_amount"]
    sheet["D13"] = result["labor_cost_amount"]
    sheet["F13"] = result["rent_cost"]

    sheet["B16"] = result["utility_cost"]
    sheet["D16"] = result["other_expense_cost"]
    sheet["F16"] = result["operating_profit"]

    sheet["B19"] = result["cost_rate"]
    sheet["D19"] = result["labor_cost_rate"]
    sheet["F19"] = result["operating_profit_rate"]

    workbook.save(output_path)

def write_store_report(
    result,
    target_month,
    store_code,
    store_name,
    template_path,
    output_path,
):
    workbook = load_workbook(template_path)
    sheet = workbook.active

    sheet["A1"] = f"【{target_month}】{store_name}月次報告書"

    sheet["B3"] = target_month
    sheet["B4"] = store_code
    sheet["D4"] = "店舗"

    sheet["B7"] = result["sales_amount_tax_ex"]
    sheet["D7"] = result["tax_amount"]
    sheet["F7"] = result["sales_amount_tax_in"]

    sheet["B10"] = result["customer_count"]
    sheet["D10"] = result["average_spend"]
    sheet["F10"] = result["tax_rate"]

    sheet["B13"] = result["cost_amount"]
    sheet["D13"] = result["labor_cost_amount"]
    sheet["F13"] = result["rent_cost"]

    sheet["B16"] = result["utility_cost"]
    sheet["D16"] = result["other_expense_cost"]
    sheet["F16"] = result["operating_profit"]

    sheet["B19"] = result["cost_rate"]
    sheet["D19"] = result["labor_cost_rate"]
    sheet["F19"] = result["operating_profit_rate"]

    workbook.save(output_path)
